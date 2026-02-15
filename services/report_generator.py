"""
Servicio de generación de reportes (JSON, PDF, Excel)
"""
import json
import os
from datetime import datetime
from typing import Dict, Any, List
from pathlib import Path

# ReportLab para PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image as RLImage
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# openpyxl para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference

from loguru import logger

from config.settings import settings
from config.database import db
from utils.helpers import (
    generate_uuid,
    sanitize_filename,
    ensure_directory,
    get_file_size,
    get_severity_color,
    format_datetime,
    JsonEncoder
)
from models.audit import ReportFormat


class ReportGeneratorService:
    """Servicio de generación de reportes"""
    
    def __init__(self):
        self.reports_dir = Path(settings.REPORTS_DIR)
        ensure_directory(self.reports_dir / "pdfs")
        ensure_directory(self.reports_dir / "excels")
        ensure_directory(self.reports_dir / "jsons")
    
    async def generate_report(
        self,
        analysis_id: str,
        audit_db_id: str,
        report_format: ReportFormat,
        analysis_data: Dict[str, Any],
        include_code: bool = True,
        include_recommendations: bool = True
    ) -> Dict[str, Any]:
        """
        Generar reporte en formato especificado
        
        Args:
            analysis_id: ID del análisis
            audit_db_id: ID de auditoría en Supabase
            report_format: Formato del reporte (JSON/PDF/EXCEL)
            analysis_data: Datos del análisis
            include_code: Incluir código en el reporte
            include_recommendations: Incluir recomendaciones
        
        Returns:
            Dict con información del reporte generado
        """
        logger.info(f"📄 Generando reporte {report_format.value} para análisis {analysis_id}")
        
        if report_format == ReportFormat.JSON:
            file_path = await self._generate_json_report(
                analysis_id, analysis_data, include_code, include_recommendations
            )
        
        elif report_format == ReportFormat.PDF:
            file_path = await self._generate_pdf_report(
                analysis_id, analysis_data, include_code, include_recommendations
            )
        
        elif report_format == ReportFormat.EXCEL:
            file_path = await self._generate_excel_report(
                analysis_id, analysis_data, include_code, include_recommendations
            )
        
        else:
            raise ValueError(f"Formato no soportado: {report_format}")
        
        # Obtener tamaño del archivo
        file_size = get_file_size(file_path)
        
        # Intentar guardar en Supabase (opcional)
        report_id = None
        try:
            report_id = await self._save_report_to_db(
                audit_db_id, report_format, file_path, file_size
            )
            logger.info(f"💾 Reporte guardado en Supabase: {report_id}")
        except Exception as e:
            logger.warning(f"⚠️ No se pudo guardar en Supabase: {e}. Continuando sin guardado.")
            report_id = analysis_id  # Usar analysis_id como fallback
        
        logger.success(f"✅ Reporte generado: {file_path}")
        
        return {
            "success": True,
            "report_id": report_id,
            "analysis_id": analysis_id,
            "report_type": report_format,
            "format": report_format.value,
            "file_path": file_path,
            "file_size": file_size,
            "download_url": f"/api/reports/download/{report_id}",
            "report_data": True  # Indicador de que el reporte fue generado
        }
    
    async def _generate_json_report(
        self,
        analysis_id: str,
        analysis_data: Dict[str, Any],
        include_code: bool,
        include_recommendations: bool
    ) -> str:
        """Generar reporte JSON"""
        
        # Preparar datos
        report_data = {
            "metadata": {
                "analysis_id": analysis_id,
                "generated_at": format_datetime(),
                "database": analysis_data.get("database_name", "Unknown"),
                "server": analysis_data.get("server", "Unknown"),
                "analyzed_sps": analysis_data.get("analyzed_count", 0),
                "total_findings": len(analysis_data.get("findings", [])),
                "generator": "AuditDB Analyzer v1.0"
            },
            "executive_summary": {
                "critical": analysis_data.get("findings_summary", {}).get("critical", 0),
                "high": analysis_data.get("findings_summary", {}).get("high", 0),
                "medium": analysis_data.get("findings_summary", {}).get("medium", 0),
                "low": analysis_data.get("findings_summary", {}).get("low", 0),
                "info": analysis_data.get("findings_summary", {}).get("info", 0),
                "risk_score": analysis_data.get("risk_score", 0.0),
                "duration_seconds": analysis_data.get("duration_seconds", 0)
            },
            "findings": []
        }
        
        # Agregar hallazgos
        for finding in analysis_data.get("findings", []):
            finding_data = {
                "id": finding["id"],
                "sp_name": finding["sp_name"],
                "category": finding["category"],
                "severity": finding["severity"],
                "type": finding["type"],
                "title": finding["title"],
                "description": finding["description"],
                "line": finding["location"]["line"],
                "impact": finding["impact"],
                "cwe_id": finding.get("cwe_id"),
                "cvss_score": finding.get("cvss_score"),
                "detected_by": finding["detected_by"]
            }
            
            if include_code:
                finding_data["code_snippet"] = finding["location"]["code_snippet"]
                finding_data["evidence"] = finding.get("evidence")
            
            if include_recommendations:
                finding_data["recommendation"] = finding["recommendation"]
            
            report_data["findings"].append(finding_data)
        
        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"auditdb_report_{analysis_id}_{timestamp}.json"
        file_path = self.reports_dir / "jsons" / filename
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False, cls=JsonEncoder)
        
        return str(file_path)
    
    async def _generate_pdf_report(
        self,
        analysis_id: str,
        analysis_data: Dict[str, Any],
        include_code: bool,
        include_recommendations: bool
    ) -> str:
        """Generar reporte PDF profesional"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"auditdb_report_{analysis_id}_{timestamp}.pdf"
        file_path = self.reports_dir / "pdfs" / filename
        
        # Crear documento
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=A4,
            rightMargin=72, leftMargin=72,
            topMargin=72, bottomMargin=18
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # ===== PORTADA =====
        story.append(Spacer(1, 2*inch))
        story.append(Paragraph("AuditDB Analyzer", title_style))
        story.append(Paragraph("Reporte de Auditoría de Base de Datos", styles['Title']))
        story.append(Spacer(1, 0.5*inch))
        
        story.append(Paragraph(
            f"<b>Fecha de Análisis:</b> {format_datetime()}",
            styles['Normal']
        ))
        story.append(Paragraph(
            f"<b>ID de Análisis:</b> {analysis_id}",
            styles['Normal']
        ))
        
        story.append(PageBreak())
        
        # ===== RESUMEN EJECUTIVO =====
        story.append(Paragraph("Resumen Ejecutivo", heading_style))
        
        summary = analysis_data.get("findings_summary", {})
        
        summary_data = [
            ['Severidad', 'Cantidad', 'Porcentaje'],
            ['🔴 CRÍTICO', str(summary.get('critical', 0)), 
             f"{self._calc_percentage(summary.get('critical', 0), len(analysis_data.get('findings', [])))}%"],
            ['🟠 ALTO', str(summary.get('high', 0)),
             f"{self._calc_percentage(summary.get('high', 0), len(analysis_data.get('findings', [])))}%"],
            ['🟡 MEDIO', str(summary.get('medium', 0)),
             f"{self._calc_percentage(summary.get('medium', 0), len(analysis_data.get('findings', [])))}%"],
            ['🔵 BAJO', str(summary.get('low', 0)),
             f"{self._calc_percentage(summary.get('low', 0), len(analysis_data.get('findings', [])))}%"],
            ['⚪ INFO', str(summary.get('info', 0)),
             f"{self._calc_percentage(summary.get('info', 0), len(analysis_data.get('findings', [])))}%"],
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 0.5*inch))
        
        # ===== HALLAZGOS CRÍTICOS =====
        critical_findings = [
            f for f in analysis_data.get("findings", [])
            if f.get("severity") == "critical"
        ]
        
        if critical_findings:
            story.append(Paragraph("Hallazgos Críticos (Acción Inmediata Requerida)", heading_style))
            
            for finding in critical_findings[:10]:  # Top 10
                story.append(Paragraph(
                    f"<b>{finding['title']}</b>",
                    styles['Heading3']
                ))
                story.append(Paragraph(
                    f"<b>SP:</b> {finding['sp_name']} | <b>Línea:</b> {finding['location']['line']}",
                    styles['Normal']
                ))
                story.append(Paragraph(
                    f"<b>Descripción:</b> {finding['description']}",
                    styles['Normal']
                ))
                
                if include_recommendations:
                    story.append(Paragraph(
                        f"<b>Recomendación:</b> {finding['recommendation']}",
                        styles['Normal']
                    ))
                
                story.append(Spacer(1, 0.2*inch))
        
        story.append(PageBreak())
        
        # ===== DETALLE DE TODOS LOS HALLAZGOS =====
        story.append(Paragraph("Detalle Completo de Hallazgos", heading_style))
        
        for finding in analysis_data.get("findings", []):
            story.append(Paragraph(
                f"<b>[{finding['severity'].upper()}] {finding['title']}</b>",
                styles['Heading4']
            ))
            story.append(Paragraph(
                f"<b>Categoría:</b> {finding['category']} | "
                f"<b>SP:</b> {finding['sp_name']} | "
                f"<b>Línea:</b> {finding['location']['line']}",
                styles['Normal']
            ))
            story.append(Paragraph(f"{finding['description']}", styles['Normal']))
            
            if include_code and finding.get('location', {}).get('code_snippet'):
                story.append(Paragraph(
                    f"<i>Código:</i> <font name=Courier>{finding['location']['code_snippet']}</font>",
                    styles['Code']
                ))
            
            story.append(Spacer(1, 0.15*inch))
        
        # Construir PDF
        doc.build(story)
        
        return str(file_path)
    
    async def _generate_excel_report(
        self,
        analysis_id: str,
        analysis_data: Dict[str, Any],
        include_code: bool,
        include_recommendations: bool
    ) -> str:
        """Generar reporte Excel con múltiples hojas"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"auditdb_report_{analysis_id}_{timestamp}.xlsx"
        file_path = self.reports_dir / "excels" / filename
        
        wb = Workbook()
        
        # ===== HOJA 1: RESUMEN EJECUTIVO =====
        ws_summary = wb.active
        ws_summary.title = "Resumen Ejecutivo"
        
        # Encabezados
        ws_summary['A1'] = "AuditDB Analyzer - Reporte de Auditoría"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        ws_summary['A3'] = "Base de Datos:"
        ws_summary['B3'] = analysis_data.get('database_name', 'N/A')
        ws_summary['A4'] = "Fecha:"
        ws_summary['B4'] = format_datetime()
        ws_summary['A5'] = "ID Análisis:"
        ws_summary['B5'] = analysis_id
        
        # Resumen de hallazgos
        ws_summary['A7'] = "Severidad"
        ws_summary['B7'] = "Cantidad"
        ws_summary['C7'] = "Porcentaje"
        
        summary = analysis_data.get("findings_summary", {})
        total_findings = len(analysis_data.get("findings", []))
        
        ws_summary['A8'] = "CRÍTICO"
        ws_summary['B8'] = summary.get('critical', 0)
        ws_summary['C8'] = f"{self._calc_percentage(summary.get('critical', 0), total_findings)}%"
        
        ws_summary['A9'] = "ALTO"
        ws_summary['B9'] = summary.get('high', 0)
        ws_summary['C9'] = f"{self._calc_percentage(summary.get('high', 0), total_findings)}%"
        
        ws_summary['A10'] = "MEDIO"
        ws_summary['B10'] = summary.get('medium', 0)
        ws_summary['C10'] = f"{self._calc_percentage(summary.get('medium', 0), total_findings)}%"
        
        ws_summary['A11'] = "BAJO"
        ws_summary['B11'] = summary.get('low', 0)
        ws_summary['C11'] = f"{self._calc_percentage(summary.get('low', 0), total_findings)}%"
        
        ws_summary['A12'] = "INFO"
        ws_summary['B12'] = summary.get('info', 0)
        ws_summary['C12'] = f"{self._calc_percentage(summary.get('info', 0), total_findings)}%"
        
        # Formatear
        self._format_excel_header(ws_summary, 'A7:C7')
        
        # ===== HOJA 2: HALLAZGOS CRÍTICOS =====
        ws_critical = wb.create_sheet("Hallazgos Críticos")
        
        critical_findings = [
            f for f in analysis_data.get("findings", [])
            if f.get("severity") == "critical"
        ]
        
        ws_critical.append(["SP", "Título", "Línea", "Descripción", "Recomendación"])
        self._format_excel_header(ws_critical, 'A1:E1')
        
        for finding in critical_findings:
            ws_critical.append([
                finding['sp_name'],
                finding['title'],
                finding['location']['line'],
                finding['description'],
                finding.get('recommendation', '') if include_recommendations else ''
            ])
        
        # ===== HOJA 3: TODOS LOS HALLAZGOS =====
        ws_all = wb.create_sheet("Todos los Hallazgos")
        
        headers = ["ID", "SP", "Severidad", "Categoría", "Título", "Línea", "Descripción"]
        if include_recommendations:
            headers.append("Recomendación")
        
        ws_all.append(headers)
        self._format_excel_header(ws_all, f'A1:{chr(64 + len(headers))}1')
        
        for finding in analysis_data.get("findings", []):
            row = [
                finding['id'],
                finding['sp_name'],
                finding['severity'].upper(),
                finding['category'],
                finding['title'],
                finding['location']['line'],
                finding['description']
            ]
            
            if include_recommendations:
                row.append(finding.get('recommendation', ''))
            
            ws_all.append(row)
        
        # ===== HOJA 4: HALLAZGOS POR CATEGORÍA =====
        ws_category = wb.create_sheet("Por Categoría")
        
        # Agrupar por categoría
        by_category = {}
        for finding in analysis_data.get("findings", []):
            cat = finding['category']
            if cat not in by_category:
                by_category[cat] = []
            by_category[cat].append(finding)
        
        ws_category.append(["Categoría", "Cantidad"])
        self._format_excel_header(ws_category, 'A1:B1')
        
        for category, findings in by_category.items():
            ws_category.append([category.upper(), len(findings)])
        
        # Guardar
        wb.save(file_path)
        
        return str(file_path)
    
    def _format_excel_header(self, worksheet, cell_range: str):
        """Formatear encabezados de Excel"""
        for row in worksheet[cell_range]:
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _calc_percentage(self, value: int, total: int) -> str:
        """Calcular porcentaje"""
        if total == 0:
            return "0.0"
        return f"{(value / total * 100):.1f}"
    
    async def _save_report_to_db(
        self,
        audit_id: str,
        report_format: ReportFormat,
        file_path: str,
        file_size: int
    ) -> str:
        """Guardar registro de reporte en Supabase"""
        report_data = {
            "audit_id": audit_id,
            "report_type": report_format.value,
            "file_path": file_path,
            "file_size": file_size
        }
        
        try:
            result = await db.insert_reporte(report_data)
            return result["id"]
        except Exception as e:
            logger.error(f"❌ Error guardando reporte en Supabase: {e}")
            raise  # Re-lanzar para que lo maneje el método que llama


# Instancia global
report_generator = ReportGeneratorService()
